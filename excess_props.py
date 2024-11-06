import copy
from py4j.java_gateway import JavaGateway
import py4j
import matplotlib.pyplot as plt
import os
import openpyxl as xl
import math

class melting_data():
    def __init__(self):
        self._wb = xl.load_workbook(
            (os.path.dirname(os.path.abspath(__file__)) + f'\\datas\\melting\\melting_data.xlsx'))
        self.melting_dict = {}
        self.pop_dict()
    def pop_dict(self):
        st = self._wb.active
        for l in filter(None,st.iter_rows(min_row=2, values_only=True)):
            if not all(value is None for value in l):
                if not f'{l[0]}' in self.melting_dict:
                    self.melting_dict[f'{l[0]}'] = {}

                self.melting_dict[f'{l[0]}']['d_Hm'] = l[1]
                self.melting_dict[f'{l[0]}']['Tm'] = l[2]


if __name__ == '__main__':
    RSI = 8.314

    selected_model = 'COSMO-SAC-HB2 (GAMESS)'
    gateway = JavaGateway(auto_field=True)
    JCOSMO = gateway.entry_point
    model = JCOSMO.newModel(selected_model)

    melting_dict = melting_data().melting_dict
    dir_datas = (os.path.dirname(os.path.abspath(__file__)) + f'\\datas')
    files = [os.path.splitext(f)[0] for f in os.listdir(dir_datas) if
             os.path.isfile(os.path.join(dir_datas, f))]
    print(files)
    dict_data = {}

    for file in files:
        nf = file

        wb = xl.load_workbook(dir_datas+f'\\{file}.xlsx')
        st = wb.active
        data = 0
        for l in filter(None, st.iter_rows(min_row=2, values_only=True)):
            if not all(value is None for value in l):
                if not f'{nf}' in dict_data:
                    dict_data[f'{nf}'] = {}
                    print(nf)
                if not f'{data}' in dict[f'{nf}']:
                    dict_data[f'{nf}'][f'{data}'] = {}
                    dict_data[f'{nf}'][f'{data}']['T'] = l[0]
                    dict_data[f'{nf}'][f'{data}']['solvent'] = l[1]
                    dict_data[f'{nf}'][f'{data}']['molar_ratio'] = l[2]
                    dict_data[f'{nf}'][f'{data}']['x_exp'] = l[3]
                    dict_data[f'{nf}'][f'{data}']['x_est'] = None
                    dict_data[f'{nf}'][f'{data}']['ln|γ|'] = None
                    dict_data[f'{nf}'][f'{data}']['ge/RT'] = None
                    dict_data[f'{nf}'][f'{data}']['he/RT'] = None
                    dict_data[f'{nf}'][f'{data}']['se/R'] = None
                T = dict_data[f'{nf}'][f'{data}']['T']
                # solute informations
                solute = nf
                d_Hm = melting_dict[solute]['d_Hm']
                Tm = melting_dict[solute]['Tm']
                sle_cte = (d_Hm / RSI) * ((1 / Tm) - (1 / T))
                # solvent informations
                solvents = l[1].split(';')
                compositon = list(map(float, str(l[2]).replace(',','.').split(';')))

                # iteration paremeter
                err = 100
                tol = pow(10, -10)
                max_iter = 200

                # start JCOSMO communication
                ncomps = len(solvents) + 1
                comps = gateway.new_array(gateway.jvm.java.lang.String, ncomps)
                x = gateway.new_array(gateway.jvm.double, ncomps)

                comps[0] = solute
                for i, solvent in enumerate(solvents):
                    comps[i+1] = solvent

                model.setCompounds(comps)
                model.setTemperature(T)
                # x[0] = 0.0
                # for i, molar_ratio in enumerate(compositon):
                #     x[i+1] = molar_ratio
                #
                # model.setComposition(x)
                # lnGamma = model.activityCoefficientLn()

                x[0] = dict_data[f'{nf}'][f'{data}']['x_exp']
                # por algum motivo, o gallic_acid em uma determinada mistura da um bug e ele nao converge, porquee
                # supoe que a solubilidade e um pouco mais alta.
                while abs(err) > tol:
                    for i, molar_ratio in enumerate(compositon):
                        x[i + 1] = molar_ratio * (1 - x[0])

                    model.setComposition(x)
                    lnGamma = model.activityCoefficientLn()

                    x_calc = math.exp(sle_cte - lnGamma[0])
                    err = x_calc - x[0]
                    x[0] = x[0] + 0.2 * err


                ge = model.excessEnthalpy()
                he = model.excessGibbs()
                dict_data[f'{nf}'][f'{data}']['x_est'] = x[0]
                dict_data[f'{nf}'][f'{data}']['ln|γ|'] = lnGamma[0]
                dict_data[f'{nf}'][f'{data}']['ge/RT'] = ge
                dict_data[f'{nf}'][f'{data}']['he/RT'] = he
                dict_data[f'{nf}'][f'{data}']['se/R'] = (he - ge) / T
                data += 1




    for k_i, v_i in dict_data.items():
        wb = xl.Workbook()
        st = wb.active
        # writing on excel file
        st['A1'] = 'Temperature'
        st['B1'] = 'Solvent'
        st['C1'] = 'Molar_ratio'
        st['D1'] = 'x_exp'
        st['E1'] = 'ln|γ|'
        st['F1'] = 'x_est'
        st['G1'] = 'ge/RT'
        st['H1'] = 'he/RT'
        st['I1'] = 'se/R'
        for k_j, v_j in v_i.items():
            k = int(k_j)
            st[f'A{k+2}'] = v_j['T']
            st[f'B{k+2}'] = v_j['solvent']
            st[f'C{k+2}'] = v_j['molar_ratio']
            st[f'D{k+2}'] = v_j['x_exp']
            st[f'E{k+2}'] = v_j['ln|γ|']
            st[f'F{k+2}'] = v_j['x_est']
            st[f'G{k+2}'] = v_j['ge/RT']
            st[f'H{k+2}'] = v_j['he/RT']
            st[f'I{k+2}'] = v_j['se/R']

        wb.save((os.path.dirname(os.path.abspath(__file__)) + f'\\output\\{k_i}.xlsx'))
