import os
from abc import ABC, abstractmethod
try:
    import zipfile
    import openpyxl
    from openpyxl.styles import Font, PatternFill
except:
    os.system('pip install openpyxl')
    import openpyxl
    from openpyxl.styles import Font, PatternFill

class WorksheetUnifac():
    def __init__(self, compound_dict, xlsx_name):
        self._font_st = Font(name='Arial',
                   size=12,
                   bold=False,
                   italic=False,
                   underline='none',
                   color='1E211E',)
        self._fill = PatternFill(fill_type='solid',
                       fgColor='6AA7E0',)
        self._workbook = openpyxl.Workbook()
        self._worksheet = self._workbook.active
        self._compound_dict = None #aqui sempre eh 3... ate segunda ordem.
        self.compoud_dict(compound_dict=compound_dict)
        self.worksheet_skull()
        self.set_worksheet()
        self.save_worksheet(xlsx_name=xlsx_name)

    def compoud_dict(self, compound_dict):
        self._compound_dict = compound_dict
    def worksheet_skull(self):
        self._worksheet['A1'] = 'Molar fraction'
        self._worksheet['B1'] = 'Compound (fix)'
        self._worksheet['C1'] = 'Compound 1'
        self._worksheet['D1'] = 'Compound 2'
        self._worksheet['A1'].font = self._font_st
        self._worksheet['B1'].font = self._font_st
        self._worksheet['C1'].font = self._font_st
        self._worksheet['D1'].font = self._font_st

    def set_worksheet(self):
        alf = ['B', 'C', 'D']
        k = 2
        # these variables below are for comparing all data
        ref_float_fraction = 9999.99
        ref_fraction = None
        for compounds_key, fractions_dict in self._compound_dict.items():
            compounds = compounds_key.split('/')
            for i, letter in enumerate(alf, start=0):
                self._worksheet[f'{letter}{k}'] = compounds[i]
                self._worksheet[f'{letter}{k}'].font = self._font_st
            k += 1

            # loop for finding the lower ln_gamma before populating the worksheet
            for fractions, lngamma in fractions_dict.items():
                if lngamma[0] < ref_float_fraction:
                    ref_float_fraction = lngamma[0]
                    ref_fraction = fractions

            for fractions, lngamma in fractions_dict.items():
                self._worksheet[f'A{k}'] = str(fractions)
                self._worksheet[f'A{k}'].font = self._font_st
                for i, letter in enumerate(alf):
                    self._worksheet[f"{letter}{k}"] = lngamma[i]
                    self._worksheet[f"{letter}{k}"].font = self._font_st
                    if fractions == ref_fraction:
                        self._worksheet[f'A{k}'].fill = self._fill
                        self._worksheet[f"{letter}{k}"].fill = self._fill
                k += 1
            ref_float_fraction = 9999.99
            ref_fraction = None

    def save_worksheet(self, xlsx_name):
        self._workbook.save(f"{xlsx_name}.xlsx")