import math
import os
from random import sample
from itertools import combinations
import zipfile
from abc import ABC, abstractmethod
try:
    from py4j.java_gateway import JavaGateway
    import zipfile
    import openpyxl
    from openpyxl.styles import Font, PatternFill
except:
    os.system('pip install py4j')
    os.system('pip install openpyxl')
    from py4j.java_gateway import JavaGateway
    import openpyxl
    from openpyxl.styles import Font, PatternFill

class ReaderJCosmo():
    def __init__(self):
        self.list_readed = []
        self.create_list()

    def create_list(self):
        dir = (os.getcwd() + '\\jcosmo3\\profiles\\HF_TZVP.zip').replace('\\ic_latos_project', '')
        with zipfile.ZipFile(dir, 'r') as file_zip:
            files_name = file_zip.namelist()
            for file_name in files_name:
                if '.gout' in file_name:
                    self.list_readed.append(file_name.replace('.gout', ''))


class GatewayJCosmo():
    def __init__(self, selected_model):
        self.gateway = JavaGateway(auto_field=True)
        self.JCOSMO = self.gateway.entry_point
        self.model = self.JCOSMO.newModel(selected_model)


class CompoundDictionary(ABC):
    def __init__(self, selected_model, temperature, solute):
        self._selected_model = selected_model
        self._temperature = temperature
        self._compounds_list = []
        self.compounds_dict = {}
        self._solute = solute
        self._reader_jcosmo = ReaderJCosmo()
        self._gatewayJCosmo = GatewayJCosmo(selected_model=self._selected_model)
        self._gateway = self._gatewayJCosmo.gateway
        self._model = self._gatewayJCosmo.model
    @property
    def compounds_list(self):
        return self._compounds_list
    @property
    def solute(self):
        return self._solute
    @property
    def temperature(self):
        return self._temperature
    @abstractmethod
    def populate_compound_list(self):
        """
        Funcao usada para popular a lista usada para o calculo
        :return:
        """
        pass
    @abstractmethod
    def indexation_dict(self, compound_1, compound_2):
        """
        funcao para criar os indexes no dicionario especifico das filhas
        :param compound_1:
        :param compound_2:
        :return:
        """
        pass

    @abstractmethod
    def indexation_lnGamma(self, compound_1, compound_2, lnGamma):
        pass
    def generate_comp_dict(self, molar_list):
        """
        funcao que gera um dicionario com lnGamma calculado
        :param molar_list:
        :return:
        """
        comps = self._gateway.new_array(self._gateway.jvm.java.lang.String, 3)
        comps[0] = self._solute

        for compound in self._compounds_list:
            self.compounds_dict[f"{self._solute}/{compound[0]}/{compound[1]}"] = {}
            self.indexation_dict(compound_1=compound[0], compound_2=compound[1])
            comps[1] = compound[0]
            comps[2] = compound[1]
            self._model.setCompounds(comps)
            self._model.setTemperature(float(self._temperature))
            x = self._gateway.new_array(self._gateway.jvm.double, 3)
            for fraction in molar_list:
                x[0] = fraction[0]
                x[1] = fraction[1]
                x[2] = fraction[2]
                self._model.setComposition(x)
                lnGamma = self._model.activityCoefficientLn()
                self.compounds_dict[f"{self._solute}/{compound[0]}/{compound[1]}"][
                    f"{fraction[0]},{fraction[1]},{fraction[2]}"] = []
                for i in range(0, 3, 1):

                    self.compounds_dict[f"{self._solute}/{compound[0]}/{compound[1]}"][
                        f"{fraction[0]},{fraction[1]},{fraction[2]}"].append(lnGamma[i])
                self.indexation_lnGamma(compound_1=compound[0], compound_2=compound[1], lnGamma=lnGamma)


class EquimolarDictComp(CompoundDictionary):
    def __init__(self, selected_model, temperature, solute, HBA_list, HBD_list):
        super().__init__(selected_model, temperature, solute)
        self.populate_compound_list(HBA_list=HBA_list, HBD_list=HBD_list)
        self._molar_list = [[0.0, 0.5, 0.5]]
        self.equimolar_comp_dict = {}
        self.generate_comp_dict(molar_list=self._molar_list)

    def populate_compound_list(self,HBA_list, HBD_list):
        for HBA in HBA_list:
            for HBD in HBD_list:
                self._compounds_list.append([HBA, HBD])

    def indexation_dict(self,compound_1, compound_2):
        if not compound_1 in self.equimolar_comp_dict:
            self.equimolar_comp_dict[compound_1] = {}

    def indexation_lnGamma(self, compound_1, compound_2, lnGamma):
        self.equimolar_comp_dict[compound_1][compound_2] = lnGamma[0]


class FixedCompDict(CompoundDictionary):
    def __init__(self, selected_model, temperature, solute, compound_1, compound_list):
        super().__init__(selected_model=selected_model, temperature=temperature, solute=solute)
        self.populate_compound_list(compound_1=compound_1, compound_list=compound_list)
        # Para alterar as fracoes molares eh so alterar em baixo
        # eh soluto / solvente_1 / solvente_2
        self._molar_list = [[0.0, 1.0, 0.0], [0.0,0.962,0.038], [0.0,0.919,0.081],
                            [0.0,0.868,0.132], [0.0,0.809,0.191],
                            [0.0,0.739,0.261], [0.0,0.653,0.347],
                            [0.0,0.548,0.452], [0.0,0.414,0.586],
                            [0.0,0.239,0.761], [0.0, 0.0, 1.0]]
        self.fixed_com_dict = {}
        self.generate_comp_dict(molar_list=self._molar_list)

    def populate_compound_list(self, compound_1, compound_list):
        for compound_2 in compound_list:
            self._compounds_list.append([compound_1, compound_2])
    def indexation_dict(self, compound_1,compound_2):
        pass

    def indexation_lnGamma(self, compound_1, compound_2, lnGamma):
        pass

class MixedCompDict(CompoundDictionary):
    def __init__(self, selected_model, temperature, solute, compound_list):
        super().__init__(selected_model=selected_model, temperature=temperature, solute=solute)
        self.populate_compound_list(compound_list=compound_list)
        # Para alterar as fracoes molares eh so alterar em baixo
        # eh soluto / solvente_1 / solvente_2
        self._molar_list = [[0.0, 0.1, 0.9], [0.0, 0.2, 0.8], [0.0, 0.3, 0.7],
                                    [0.0, 0.4, 0.6], [0.0, 0.5, 0.5],[0.0, 0.6, 0.4],
                                    [0.0, 0.7, 0.3],[0.0, 0.8, 0.2], [0.0, 0.9, 0.1]]
        self.fixed_com_dict = {}
        self.generate_comp_dict(molar_list=self._molar_list)

    def populate_compound_list(self, compound_list):
        self._compounds_list = list(combinations(compound_list, 2))
    def indexation_dict(self, compound_1,compound_2):
        pass
    def indexation_lnGamma(self, compound_1, compound_2, lnGamma):
        pass

class MonoComp():
    def __init__(self, selected_model, temp_f, temp_o, N, solute, solvent):
        self._selected_model = selected_model
        _h = (temp_f - temp_o) / N
        self.list_temperature = []
        for T in range(N+1):
            self.list_temperature.append(float(temp_o + T*_h))
        self._solute = solute
        self._solvent = solvent
        self._molar_list = [0.0, 1.0]
        self.compounds_dict = {}
        self._reader_jcosmo = ReaderJCosmo()
        self._gatewayJCosmo = GatewayJCosmo(selected_model=self._selected_model)
        self._gateway = self._gatewayJCosmo.gateway
        self._model = self._gatewayJCosmo.model
        self.populate_compound_list()

    def populate_compound_list(self):
        comps = self._gateway.new_array(self._gateway.jvm.java.lang.String, 2)
        comps[0] = self._solute

        x = self._gateway.new_array(self._gateway.jvm.double, 2)
        x[0] = self._molar_list[0]
        x[1] = self._molar_list[1]

        for solvent in self._solvent:
            if not solvent in self.compounds_dict:
                self.compounds_dict[solvent] = {}
            comps[1] = solvent
            self._model.setCompounds(comps)
            self._model.setComposition(x)
            for temperature in self.list_temperature:
                self._model.setTemperature(float(temperature))
                lnGamma = self._model.activityCoefficientLn()
                self.compounds_dict[solvent][temperature] = lnGamma[0] #MUDAR AQUI



class Save_MonoComp: #VOU TER QUE MUDAR
    def __init__(self, compound_dict, xlsx_name):
        self._font_st = Font(name='Arial',
                             size=12,
                             bold=False,
                             italic=False,
                             underline='none',
                             color='1E211E', )
        self._fill = PatternFill(fill_type='solid',
                                 fgColor='6AA7E0', )
        self._workbook = openpyxl.Workbook()
        self._worksheet = self._workbook.active
        self._compound_dict = None  # aqui sempre eh 3... ate segunda ordem.
        self.compoud_dict(compound_dict=compound_dict)
        self._qtd_compounds = len(set(self._compound_dict.keys()))
        self._alf = [chr(ord('B') + i) for i in range(self._qtd_compounds)]
        self.worksheet_skull()
        self.set_worksheet()
        self.save_worksheet(xlsx_name=xlsx_name)


    def compoud_dict(self, compound_dict):
        self._compound_dict = compound_dict
    def worksheet_skull(self):
        self._worksheet['A1'] = 'Temperature'
        self._worksheet['A1'].font = self._font_st
        for i, (solvent, values) in enumerate(self._compound_dict.items()):
            self._worksheet[f'{self._alf[i]}1'] = solvent
            self._worksheet[f'{self._alf[i]}1'].font = self._font_st

    def set_worksheet(self):
        for i, (solvent, temperatures) in enumerate(self._compound_dict.items()):
            k = 2
            for temperature, lnGamma in temperatures.items():
                self._worksheet[f'A{k}'] = temperature
                self._worksheet[f'{self._alf[i]}{k}'] = lnGamma
                k += 1

    def save_worksheet(self, xlsx_name):
        self._workbook.save(f"{xlsx_name}.xlsx")

class Worksheet():
    def __init__(self, compound_dict, xlsx_name):
        self._font_st = Font(name='Arial',
                   size=12,
                   bold=False,
                   italic=False,
                   underline='none',
                   color='1E211E',)
        self._fill = PatternFill(fill_type='solid',
                       fgColor='6AA7E0',)
        self._workbook = openpyxl.Workbook()
        self._worksheet = self._workbook.active
        self._compound_dict = None #aqui sempre eh 3... ate segunda ordem.
        self.compoud_dict(compound_dict=compound_dict)
        self.worksheet_skull()
        self.set_worksheet()
        self.save_worksheet(xlsx_name=xlsx_name)

    def compoud_dict(self, compound_dict):
        self._compound_dict = compound_dict
    def worksheet_skull(self):
        self._worksheet['A1'] = 'Molar fraction'
        self._worksheet['B1'] = 'Compound (fix)'
        self._worksheet['C1'] = 'Compound 1'
        self._worksheet['D1'] = 'Compound 2'
        self._worksheet['A1'].font = self._font_st
        self._worksheet['B1'].font = self._font_st
        self._worksheet['C1'].font = self._font_st
        self._worksheet['D1'].font = self._font_st

    def set_worksheet(self):
        alf = ['B', 'C', 'D']
        k = 2
        # these variables below are for comparing all data
        ref_float_fraction = 9999.99
        ref_fraction = None
        for compounds_key, fractions_dict in self._compound_dict.items():
            compounds = compounds_key.split('/')
            for i, letter in enumerate(alf, start=0):
                self._worksheet[f'{letter}{k}'] = compounds[i]
                self._worksheet[f'{letter}{k}'].font = self._font_st
            k += 1

            # loop for finding the lower ln_gamma before populating the worksheet
            for fractions, lngamma in fractions_dict.items():
                if lngamma[0] < ref_float_fraction:
                    ref_float_fraction = lngamma[0]
                    ref_fraction = fractions

            for fractions, lngamma in fractions_dict.items():
                self._worksheet[f'A{k}'] = str(fractions)
                self._worksheet[f'A{k}'].font = self._font_st
                for i, letter in enumerate(alf):
                    self._worksheet[f"{letter}{k}"] = lngamma[i]
                    self._worksheet[f"{letter}{k}"].font = self._font_st
                    if fractions == ref_fraction:
                        self._worksheet[f'A{k}'].fill = self._fill
                        self._worksheet[f"{letter}{k}"].fill = self._fill
                k += 1
            ref_float_fraction = 9999.99
            ref_fraction = None

    def save_worksheet(self, xlsx_name):
        self._workbook.save(f"{xlsx_name}.xlsx")

class Partition_K():
    def __init__(self, selected_model, solutes, solvents, temperature):
        self._selected_model = selected_model
        self._solutes = solutes
        self._solvents = solvents # recebe como [['Name', Total Concentration, Density, MW], ['1-OCTANOL', 8.37, 0.824, 130.2279]]
        self._dict_solvents = {}
        self._temperature = temperature
        self._molar_list = [0.0, 1.0]
        self.dict_lnGamma = {}
        self.dict_K = {}

        self._reader_jcosmo = ReaderJCosmo()
        self._gatewayJCosmo = GatewayJCosmo(selected_model=self._selected_model)
        self._gateway = self._gatewayJCosmo.gateway
        self._model = self._gatewayJCosmo.model

        self.dict_solv_constructor()
        self.create_dict_lnGamma()
        self.create_dict_K()

    def dict_solv_constructor(self):
        i = 1
        for solvent in self._solvents:
            self._dict_solvents[f'{i}'] = {}
            self._dict_solvents[f'{i}']['Name'] = solvent[0]
            self._dict_solvents[f'{i}']['Total concentration'] = solvent[1]
            self._dict_solvents[f'{i}']['Density'] = solvent[2]
            self._dict_solvents[f'{i}']['MW'] = solvent[3]
            i += 1

    def create_dict_lnGamma(self):
        comps = self._gateway.new_array(self._gateway.jvm.java.lang.String, 2)
        x = self._gateway.new_array(self._gateway.jvm.double, 2)

        for key, value in self._dict_solvents.items():
            comps[1] = value['Name']

            for solute in self._solutes:
                if not solute in self.dict_lnGamma:
                    self.dict_lnGamma[solute] = {}
                comps[0] = solute

                x[0] = self._molar_list[0]
                x[1] = self._molar_list[1]

                self._model.setCompounds(comps)
                self._model.setComposition(x)
                self._model.setTemperature(self._temperature)
                lnGamma = self._model.activityCoefficientLn()
                self.dict_lnGamma[solute][value['Name']] = math.exp(lnGamma[0])

    def create_dict_K(self):
        solvent_1 = self._dict_solvents['1']['Name']
        solvent_2 = self._dict_solvents['2']['Name']

        C_C_1 = self._dict_solvents['2']['Total concentration'] / self._dict_solvents['1']['Total concentration']
        C_C_2 = (self._dict_solvents['2']['Density'] / self._dict_solvents['2']['MW']) / (
                    self._dict_solvents['1']['Density'] / self._dict_solvents['1']['MW'])

        for solute, value in self.dict_lnGamma.items():
            if not solute in self.dict_K:
                self.dict_K[solute] = {}
            self.dict_K[solute]['K_1'] = math.log10(C_C_1*(value[solvent_1] / value[solvent_2]))
            self.dict_K[solute]['K_2'] = math.log10(C_C_2*(value[solvent_1] / value[solvent_2]))
            self.dict_K[solute]['K_3'] = math.log10(value[solvent_1] / value[solvent_2])



class Save_partition_K():
    def __init__(self, dict_K, xlsx_name):
        self._font_st = Font(name='Arial',
                   size=12,
                   bold=False,
                   italic=False,
                   underline='none',
                   color='1E211E',)
        self._fill = PatternFill(fill_type='solid',
                       fgColor='6AA7E0',)
        self._workbook = openpyxl.Workbook()
        self._worksheet = self._workbook.active
        self._dict_K = None #aqui sempre eh 3... ate segunda ordem.
        self.dict_K(dict_K)
        self.worksheet_skull()
        self.set_worksheet()
        self.save_worksheet(xlsx_name=xlsx_name)

    def dict_K(self, dict_K):
        self._dict_K = dict_K

    def worksheet_skull(self):
        self._worksheet['A1'] = 'Compound'
        self._worksheet['B1'] = 'log K_1 (CC * gamma_1/gamma_2)'
        self._worksheet['C1'] = 'log K_2 (CC_rough * gamma_1/gamma_2)'
        self._worksheet['D1'] = 'log K_3 (gamma_2/gamma_1)'
        self._worksheet['A1'].font = self._font_st
        self._worksheet['B1'].font = self._font_st
        self._worksheet['C1'].font = self._font_st
        self._worksheet['D1'].font = self._font_st

    def set_worksheet(self):
        k = 2
        for solute, value_K in self._dict_K.items():
            self._worksheet[f'A{k}'] = solute
            self._worksheet[f'B{k}'] = value_K['K_1']
            self._worksheet[f'C{k}'] = value_K['K_2']
            self._worksheet[f'D{k}'] = value_K['K_3']
            k += 1

    def save_worksheet(self, xlsx_name):
        self._workbook.save(f"{xlsx_name}.xlsx")






class SigmaProfile():
    def __init__(self, selected_model, compound, temperature):
        self._selected_model = selected_model
        self._compound = compound
        self._temperature = temperature
        self._molar_list = [1.0]
        self._reader_jcosmo = ReaderJCosmo()
        self._gatewayJCosmo = GatewayJCosmo(selected_model=self._selected_model)
        self._gateway = self._gatewayJCosmo.gateway
        self._model = self._gatewayJCosmo.model
        self.teste()

    def teste(self):
        comps = self._gateway.new_array(self._gateway.jvm.java.lang.String, 1)
        x = self._gateway.new_array(self._gateway.jvm.double, 1)
        comps[0] = self._compound
        x[0] = self._molar_list[0]

        self._model.setCompounds(comps)
        self._model.setComposition(x)
        self._model.setTemperature(self._temperature)
        sigma_HB = self._model.getSigmaHB()
        print(sigma_HB)

        beta = self._model.getBeta()
        print(beta)

        density = self._model.excessGibbs()
        density2 = self._model.activityCoefficientLn()
        print(density)
        print(density2[0])
        print(self._model.help())





#daqui pra baixo to para apagar
class CompDict_x():
    def __init__(self, selected_model, list_type, comp_0, temperature, n_comp_sol):
        self.selected_model = selected_model
        self._list_type = None
        self.comp_0 = comp_0
        self.n_comp_sol = n_comp_sol
        self.temperature = temperature
        """self.list_molar_fraction = [[0.0, 0.1, 0.9], [0.0, 0.2, 0.8], [0.0, 0.3, 0.7],
                                   [0.0, 0.4, 0.6], [0.0, 0.5, 0.5],
                                   [0.0, 0.6, 0.4], [0.0, 0.7, 0.3],
                                   [0.0, 0.8, 0.2], [0.0, 0.9, 0.1]]"""
        self.list_molar_fraction = [[0.0,0.5,0.5]]
        self._reader_jcosmo = ReaderJCosmo()
        self._gateJCosmo = GatewayJCosmo(selected_model=self.selected_model)
        self._gateway = self._gateJCosmo.gateway
        self._model = self._gateJCosmo.model
        self._compounds_list = []
        self.compounds_dict = {}

        self.list_type(list_type)
        self.create_list_compounds()
        self.create_dict()
        if list_type == 2 and self.list_molar_fraction == [[0.0,0.5,0.5]]:
            self.equimolar_dict = {}
            self.teste_create()

    def list_type(self, list_type):
        self._list_type = list_type

    def random_list(self, n_list_compounds):
        """
        eh para criar uma lista valida de todos os components, posteriormente podemos pegar numa
        gui que facilite quem manusear o sistema a usá-lo. Deixar bastante semelhante ao JCOSMO.
        :param n_list_compounds:
        :return:
        """
        list_readed = self._reader_jcosmo.list_readed
        list_to_create = self._gateway.new_array(self._gateway.jvm.java.lang.String, n_list_compounds)
        list_compouds = sample(list_readed, n_list_compounds)

        while True:
            try:
                for i, compound in enumerate(list_compouds):
                    list_to_create[i] = compound
                self._model.setCompounds(list_to_create)
                break
            except:
                list_compouds = sample(list_readed, n_list_compounds)
        return list_compouds

    def create_list_compounds(self):
        if self._list_type == 1:
            l_compounds = [
                'TERT-BUTYLCYCLOHEXANE',
                'O-TOLUNITRILE',
                'ISOBUTYRALDEHYDE',
                '3-METHYL-1-PENTYNE',
                '2-ETHYL-6-METHYLNAPHTHALENE',
            ]
            self._compounds_list = list(combinations(l_compounds, 2))
        elif self._list_type == 2:
            HBA_compounds = ['PROLINE',
                            'BETAINE',
                            'DGLUCOSE',
                            'GLUCOSE',
                            'FRUCTOSE',
                            'SUCROSE',
                            'MALTOSE',
                            'XYLOSE',
                            'CITRIC_ACID',
                            'LACTIC_ACID',
                            'MALIC_ACID',
                            'L-MENTHOL',
                            'GLYCEROL',
                               ]


            HBD_compounds = ['FORMIC_ACID',
                            'ACETIC_ACID',
                            'OXALIC_ACID',
                            'MALONIC_ACID',
                            'TARTARIC_ACID',
                            'LEVULINIC_ACID',
                            'PROPIONIC_ACID',
                            'N-HEXANOIC_ACID',
                            'N-OCTANOIC_ACID',
                            'N-DECANOIC_ACID',
                            'N-DODECANOIC_ACID',
                            'N-TETRADECANOIC_ACID',
                            'OLEIC_ACID',
                            'RICINOLEIC_ACID',
                            'BENZOIC_ACID',
                            '4-HYDROXY-PHENYLACETIC_ACID',
                            '1-HEXANOL',
                            '1-OCTANOL',
                            '1-DECANOL',
                            '1-DODECANOL',
                            '1-TETRADECANOL',
                            '1-HEXADECANOL',
                            'ETHYLENE_GLYCOL',
                            'TRIETHYLENE_GLYCOL',
                            '1,3-PROPANEDIOL',
                            '1,2-BUTANEDIOL',
                            '1,3-BUTANEDIOL',
                            '1,4-BUTANEDIOL',
                            '1,6-HEXANEDIOL',
                            'CYCLOHEXANOL',
                            'UREA',
                            ]

            for HBA in HBA_compounds:
                for HBD in HBD_compounds:
                    self._compounds_list.append([HBA, HBD])
        else:
            l_compounds = self.random_list(n_list_compounds=10)
            self._compounds_list = list(combinations(l_compounds, 2))


    def create_dict(self):
        comps = self._gateway.new_array(self._gateway.jvm.java.lang.String, self.n_comp_sol)
        comps[0] = self.comp_0

        for compound in self._compounds_list:
            comps[1] = compound[0]
            comps[2] = compound[1]

            self.compounds_dict[f"{self.comp_0}/{compound[0]}/{compound[1]}"] = {}

            self._model.setCompounds(comps)
            self._model.setTemperature(float(self.temperature))
            x = self._gateway.new_array(self._gateway.jvm.double, self.n_comp_sol)
            for fraction in self.list_molar_fraction:
                x[0] = fraction[0]
                x[1] = fraction[1]
                x[2] = fraction[2]
                self._model.setComposition(x)
                lnGamma = self._model.activityCoefficientLn()
                self.compounds_dict[f"{self.comp_0}/{compound[0]}/{compound[1]}"][
                    f"{fraction[0]},{fraction[1]},{fraction[2]}"] = []
                for i in range(0, 3, 1):
                    self.compounds_dict[f"{self.comp_0}/{compound[0]}/{compound[1]}"][
                        f"{fraction[0]},{fraction[1]},{fraction[2]}"].append(lnGamma[i])

    def teste_create(self):
        """
        aqui preciso fazer uma lista de 2 a 2, porque ela devera ser usada para testar acidos e tals
        :return:
        """
        comps = self._gateway.new_array(self._gateway.jvm.java.lang.String, self.n_comp_sol)
        comps[0] = self.comp_0

        for compound in self._compounds_list:
            if not compound[0] in self.equimolar_dict:
                self.equimolar_dict[compound[0]] = {}
            comps[1] = compound[0]
            comps[2] = compound[1]
            self._model.setCompounds(comps)
            self._model.setTemperature(float(self.temperature))
            x = self._gateway.new_array(self._gateway.jvm.double, self.n_comp_sol)
            x[0] = self.list_molar_fraction[0][0]
            x[1] = self.list_molar_fraction[0][1]
            x[2] = self.list_molar_fraction[0][2]
            self._model.setComposition(x)
            lnGamma = self._model.activityCoefficientLn()
            self.equimolar_dict[compound[0]][compound[1]] = lnGamma[0]



