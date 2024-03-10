import os
import openpyxl
from openpyxl.styles import Font, PatternFill
import  numpy as np


def encontrar_indice(value, dict):
    for k, v in dict.items():
        if value in v:
            return k
    return None


class SheetsOpener():
    def __init__(self):
        dir_main = (os.getcwd() + '\\models\\sheets\\unifac_main.xlsx')
        dir_sub = (os.getcwd() + '\\models\\sheets\\unifac_sub.xlsx')
        self.main_sheet = openpyxl.load_workbook(dir_main)['main']
        self.sub_sheet = openpyxl.load_workbook(dir_sub)['sub']


# modelo retirado do apendice G van ness

class Comp():
    def __init__(self, name, x, list_group):
        self._sheets_opener = SheetsOpener()
        self.comp = {
            'name': name,
            'x': x,
            'r': 0.0,
            'q': 0.0,
            'J': 0.0,
            'L': 0.0,
            'ln_comb': 0.0,
            'ln_res': 0.0,
            'ln': 0.0,
            'group': None
        }
        self.populate_group(list_group=list_group)

    def __getitem__(self, key):
        return self.comp[key]

    def __setitem__(self, key, value):
        self.comp[key] = value

    def populate_group(self,list_group):
        _dict_aux = {}
        for i in list_group:
            for num_row, row in enumerate(
                    self._sheets_opener.main_sheet.iter_rows(min_row=1,
                                                             max_row=self._sheets_opener.main_sheet.max_row,
                                                             min_col=1,
                                                             max_col=1,
                                                             values_only=True)):
                if i[0] == row[0]:
                    main_group = self._sheets_opener.main_sheet.cell(row=num_row + 1, column=3).value
                    sub_group = self._sheets_opener.main_sheet.cell(row=num_row + 1, column=1).value
                    if not main_group in _dict_aux:
                        _dict_aux[main_group] = { }
                    _R = self._sheets_opener.main_sheet.cell(row=num_row + 1, column=4).value
                    _Q = self._sheets_opener.main_sheet.cell(row=num_row + 1, column=5).value
                    _dict_aux[main_group][sub_group] = {
                        'qtd': i[1],
                        'R': _R,
                        'Q': _Q
                    }
        self.comp['group'] = _dict_aux


class UNIFAC():
    def __init__(self, temp_f, temp_o, n, list_comp):
        self._sheets_opener = SheetsOpener()
        self.comps = {}
        self.temp_f = temp_f
        self.temp_o = temp_o
        self.N = n

        self._soma_R_comps = 0.0
        self._soma_Q_comps = 0.0

        for i, comp in enumerate(list_comp):
            self.comps[f'comp_{i}'] = comp
        self.string_dict = ""

        for i_comp, v_comp in self.comps.items():
            self.string_dict += v_comp['name'] + ';'
        self.string_dict = self.string_dict.rstrip(';')

        self.TE = {self.string_dict: {}
                   }

        self.populate_dict()
        self.ln_gamma_comb()
        self.ln_gamma_res()

    def populate_dict(self):
        for k_comp, v_comp in self.comps.items():
            soma_R = 0.0
            soma_Q = 0.0
            for k_main, v_main in v_comp['group'].items():
                for k_group, v_group in v_main.items():
                    for row in self._sheets_opener.main_sheet.iter_rows(min_row=2,
                                                                        max_row=self._sheets_opener.main_sheet.max_row,
                                                                        min_col=1,
                                                                        max_col=1,
                                                                        values_only=True):
                        if k_group in row:
                            v_group['R'] = self._sheets_opener.main_sheet.cell(row=row[0] + 1, column=4).value
                            v_group['Q'] = self._sheets_opener.main_sheet.cell(row=row[0] + 1, column=5).value
                            soma_R += (v_group['R'] * v_group['qtd'])
                            soma_Q += (v_group['Q'] * v_group['qtd'])
            v_comp['r'] = soma_R
            v_comp['q'] = soma_Q

    def ln_gamma_comb(self):
        self._soma_R_comps = 0.0
        self._soma_Q_comps = 0.0
        for k_comp, v_comp in self.comps.items():
            self._soma_R_comps += v_comp['x'] * v_comp['r']
            self._soma_Q_comps += v_comp['x'] * v_comp['q']

        for k_comp, v_comp in self.comps.items():
            v_comp['J'] = v_comp['r'] / self._soma_R_comps
            v_comp['L'] = v_comp['q'] / self._soma_Q_comps
            _J = v_comp['J']
            _L = v_comp['L']
            _Q = v_comp['q']
            v_comp['ln_comb'] = 1 - _J + np.log(_J) - 5 * _Q * (1 - (_J / _L) + np.log(_J / _L))

    def ln_gamma_res(self):
        unique_groups = set()
        unique_subgroups = set()
        dict_aux = {}
        for k_comp, v_comp in self.comps.items():
            for k_main, v_main in v_comp['group'].items():
                unique_groups.add(k_main)
                if not k_main in dict_aux:
                    dict_aux[k_main] = []
                for k_group in v_main:
                    unique_subgroups.add(k_group)
                    if not k_group in dict_aux[k_main]:
                        dict_aux[k_main].append(k_group)

        orded_subgroup = sorted(list(unique_subgroups))
        orded_group = sorted(list(unique_groups))
        orded_dict = dict(sorted(dict_aux.items()))
        list_aux = []
        for k_i, v_i in orded_dict.items():
            for i in v_i:
                for k_j, v_j in orded_dict.items():
                    for j in v_j:
                        list_aux.append([encontrar_indice(i, orded_dict), encontrar_indice(j, orded_dict)])

        a_mk = np.zeros((len(orded_subgroup), len(orded_subgroup)), dtype=object)
        for i in range(len(list_aux)):
            r = i // len(orded_subgroup)
            c = i % len(orded_subgroup)
            a_mk[r][c] = list_aux[i]

        # gera a matriz de coeficiente combinatorios
        for i in range(len(orded_subgroup)):
            for j in range(len(orded_subgroup)):
                if a_mk[i][j][0] == a_mk[i][j][1]:
                    a_mk[i][j] = 0
                else:
                    if a_mk[i][j][0] < a_mk[i][j][1]:
                        for num_row, row in enumerate(
                                self._sheets_opener.sub_sheet.iter_rows(min_row=self._sheets_opener.sub_sheet.min_row,
                                                                        max_row=self._sheets_opener.sub_sheet.max_row,
                                                                        min_col=1,
                                                                        max_col=2,
                                                                        values_only=True)):
                            if row[0] == a_mk[i][j][0] and row[1] == a_mk[i][j][1]:
                                v = self._sheets_opener.sub_sheet.cell(row=num_row + 1, column=3).value

                        a_mk[i][j] = v
                    else:
                        for num_row, row in enumerate(
                                self._sheets_opener.sub_sheet.iter_rows(min_row=self._sheets_opener.sub_sheet.min_row,
                                                                        max_row=self._sheets_opener.sub_sheet.max_row,
                                                                        min_col=1,
                                                                        max_col=2,
                                                                        values_only=True)):
                            if row[0] == a_mk[i][j][1] and row[1] == a_mk[i][j][0]:
                                v = self._sheets_opener.sub_sheet.cell(row=num_row + 1, column=4).value
                        a_mk[i][j] = v

        e_ki = np.zeros((len(orded_subgroup), len(self.comps)))

        list_aux = np.zeros((len(self.comps), len(orded_subgroup)), dtype=object)

        for num_row, ind in enumerate(self.comps.items()):
            for k_main, v_main in ind[1]['group'].items():
                for k_group, v_group in v_main.items():
                    i = orded_subgroup.index(k_group)
                    list_aux[num_row][i] = v_group['qtd'] * v_group['Q'] / ind[1]['q']

        for num_col, col in enumerate(list_aux):
            for num_row, row in enumerate(col):
                e_ki[num_row][num_col] = row

        # theta
        theta = []
        list_aux = []
        for k_comp, v_comp in self.comps.items():
            list_aux.append(v_comp['x'] * v_comp['q'])
        for i in range(len(unique_subgroups)):
            soma = 0
            for j in range(len(self.comps)):
                soma += e_ki[i][j] * list_aux[j]
            theta.append(soma / self._soma_Q_comps)

        # Temperature dependency
        h = (self.temp_f - self.temp_o) / self.N
        for t in range(self.N + 1):
            _T = (self.temp_o + h * t)
            if _T not in self.TE[self.string_dict]:
                self.TE[self.string_dict][_T] = []

            tal_mk = np.zeros((len(orded_subgroup), len(orded_subgroup)))
            for num_row, row in enumerate(a_mk):
                for num_col, col in enumerate(row):
                    tal_mk[num_row][num_col] = np.exp((-col / _T))
            # beta
            beta_ik = np.zeros((len(self.comps), len(orded_subgroup)))

            for i in range(len(self.comps)):
                for k in range(len(unique_subgroups)):
                    soma = 0
                    for m in range(len(unique_subgroups)):
                        soma += e_ki[m][i] * tal_mk[m][k]
                    beta_ik[i][k] = soma

            # sk
            s_k = []
            for i in range(len(unique_subgroups)):
                soma = 0
                for j in range(len(unique_subgroups)):
                    soma += theta[j] * tal_mk[j][i]
                s_k.append(soma)

            # calcular o ln de gamma residual
            for i_comp, v_comp in enumerate(self.comps.items()):
                soma = 0
                for k in range(len(unique_subgroups)):
                    soma += theta[k] * beta_ik[i_comp][k] / s_k[k] - e_ki[k][i_comp] * np.log(
                        beta_ik[i_comp][k] / s_k[k])
                v_comp[1]['ln_res'] = v_comp[1]['q'] * (1 - soma)
                v_comp[1]['ln'] = np.exp(v_comp[1]['ln_res'] + v_comp[1]['ln_comb'])
                self.TE[self.string_dict][_T].append(v_comp[1]['ln'])


"""
dir_main = (os.getcwd() + '\\sheets\\unifac_main.xlsx')
dir_sub = (os.getcwd() + '\\sheets\\unifac_sub.xlsx')
main_sheet = openpyxl.load_workbook(dir_main)['main']
sub_sheet = openpyxl.load_workbook(dir_sub)['sub']


def monta_comp(name,x,list_group):
    comp = {
        'name': name,
        'x': x,
        'r': 0.0,
        'q': 0.0,
        'J': 0.0,
        'L': 0.0,
        'ln_comb': 0.0,
        'ln_res': 0.0,
        'ln': 0.0,
        'group': None

    }

    dict_aux = {}
    for i in list_group:
        for num_row, row in enumerate(
                main_sheet.iter_rows(min_row=1, max_row=main_sheet.max_row, min_col=1, max_col=1, values_only=True)):
            if i[0] == row[0]:
                main_group = main_sheet.cell(row=num_row + 1, column=3).value
                sub_group = main_sheet.cell(row=num_row + 1, column=1).value
                if not main_group in dict_aux:
                    dict_aux[main_group] = {}
                R = main_sheet.cell(row=num_row + 1, column=4).value
                Q = main_sheet.cell(row=num_row + 1, column=5).value
                dict_aux[main_group][sub_group] = {
                    'qtd':i[1],
                    'R':R,
                    'Q':Q
                }
    comp['group'] = dict_aux

    return comp"""


# VENDO DE DELETAR
"""# popular dicionarios
for k_comp, v_comp in comps.items():
    soma_R = 0.0
    soma_Q = 0.0
    for k_main, v_main in v_comp['group'].items():
        for k_group, v_group in v_main.items():
            for row in main_sheet.iter_rows(min_row=2, max_row=main_sheet.max_row, min_col=1, max_col=1,values_only=True):
                if k_group in row:
                    v_group['R'] = main_sheet.cell(row=row[0] + 1, column=4).value
                    v_group['Q'] = main_sheet.cell(row=row[0] + 1, column=5).value
                    soma_R += (v_group['R'] * v_group['qtd'])
                    soma_Q += (v_group['Q'] * v_group['qtd'])
    v_comp['r'] = soma_R
    v_comp['q'] = soma_Q

# ln gamma combinatorio (ate aqui ok)
soma_R_comps = 0.0
soma_Q_comps = 0.0
for k_comp, v_comp in comps.items():
    soma_R_comps += v_comp['x'] * v_comp['r']
    soma_Q_comps += v_comp['x'] * v_comp['q']
for k_comp, v_comp in comps.items():
    v_comp['J'] = v_comp['r'] / soma_R_comps
    v_comp['L'] = v_comp['q'] / soma_Q_comps
    J = v_comp['J']
    L = v_comp['L']
    Q = v_comp['q']
    v_comp['ln_comb'] = 1 - J + np.log(J) - 5*Q*(1 - (J/L)+np.log(J/L))

## variaveis para determianr ln_residual
grupos_unicos = set()
subgrupos_unicos = set()
dict_aux = {}
for k_comp, v_comp in comps.items():
    for k_main, v_main in v_comp['group'].items():
        grupos_unicos.add(k_main)
        if not k_main in dict_aux:
            dict_aux[k_main] = []
        for k_group in v_main:
            subgrupos_unicos.add(k_group)
            if not k_group in dict_aux[k_main]:
                dict_aux[k_main].append(k_group)

orded_subgroup = sorted(list(subgrupos_unicos))
orded_group = sorted(list(grupos_unicos))
orded_dict = dict(sorted(dict_aux.items()))
list_aux = []
for k_i, v_i in orded_dict.items():
    for i in v_i:
        for k_j, v_j in orded_dict.items():
            for j in v_j:
                list_aux.append([encontrar_indice(i, orded_dict), encontrar_indice(j, orded_dict)])

a_mk = np.zeros((len(orded_subgroup), len(orded_subgroup)),dtype=object)
for i in range(len(list_aux)):
    r = i // len(orded_subgroup)
    c = i % len(orded_subgroup)
    a_mk[r][c] = list_aux[i]

## gera a matriz de coeficiente combinatorios
for i in range(len(orded_subgroup)):
    for j in range(len(orded_subgroup)):
        if a_mk[i][j][0] == a_mk[i][j][1]:
            a_mk[i][j] = 0
        else:
            if a_mk[i][j][0] < a_mk[i][j][1]:
                for num_row, row in enumerate(sub_sheet.iter_rows(min_row=sub_sheet.min_row, max_row=sub_sheet.max_row, min_col=1, max_col=2,
                                                values_only=True)):
                    if row[0] == a_mk[i][j][0] and row[1] == a_mk[i][j][1]:
                        v = sub_sheet.cell(row=num_row + 1, column=3).value

                a_mk[i][j] = v
            else:
                for num_row, row in enumerate(sub_sheet.iter_rows(min_row=sub_sheet.min_row, max_row=sub_sheet.max_row, min_col=1, max_col=2,
                                                values_only=True)):
                    if row[0] == a_mk[i][j][1] and row[1] == a_mk[i][j][0]:
                        v = sub_sheet.cell(row=num_row + 1, column=4).value
                a_mk[i][j] = v

e_ki = np.zeros((len(orded_subgroup), len(comps)))

list_aux = np.zeros((len(comps), len(orded_subgroup)),dtype=object)

for num_row, ind in enumerate(comps.items()):
    for k_main, v_main in ind[1]['group'].items():
        for k_group, v_group in v_main.items():
            i = orded_subgroup.index(k_group)
            list_aux[num_row][i] = v_group['qtd']*v_group['Q'] / ind[1]['q']

for num_col, col in enumerate(list_aux):
    for num_row, row in enumerate(col):
        e_ki[num_row][num_col] = row


# theta
theta = []
list_aux = []
for k_comp, v_comp in comps.items():
    list_aux.append(v_comp['x']*v_comp['q'])
for i in range(len(subgrupos_unicos)):
    soma = 0
    for j in range(len(comps)):
        soma += e_ki[i][j]*list_aux[j]
    theta.append(soma/soma_Q_comps)


string_dict = ""
for i_comp, v_comp in comps.items():
    string_dict += v_comp['name'] + ';'
string_dict = string_dict.rstrip(';')

TE = {string_dict:
          { }
      }

print(TE)

## O QUE DEPENDE DA TEMPERATURA
Ti = 307
Tf = 325
N = 5
h = (Tf - Ti) / N
dict_T = {}
for i in range(N+1):
    T = (Ti +h*i)
    if not T in dict_T:
        dict_T[T] = []

    if not T in TE[string_dict]:
        TE[string_dict][T] = []

    tal_mk = np.zeros((len(orded_subgroup), len(orded_subgroup)))
    for num_row, row in enumerate(a_mk):
        for num_col, col in enumerate(row):
            tal_mk[num_row][num_col] = np.exp((-col/T))
    # beta
    beta_ik = np.zeros((len(comps), len(orded_subgroup)))

    for i in range(len(comps)):
        for k in range(len(subgrupos_unicos)):
            soma = 0
            for m in range(len(subgrupos_unicos)):
                soma += e_ki[m][i]*tal_mk[m][k]
            beta_ik[i][k] = soma

    # sk
    s_k = []
    for i in range(len(subgrupos_unicos)):
        soma = 0
        for j in range(len(subgrupos_unicos)):
            soma += theta[j]*tal_mk[j][i]
        s_k.append(soma)
    # calcular o ln de gamma residual
    for i_comp, v_comp in enumerate(comps.items()):
        soma = 0
        for k in range(len(subgrupos_unicos)):
            soma += theta[k]*beta_ik[i_comp][k]/s_k[k]-e_ki[k][i_comp]*np.log(beta_ik[i_comp][k]/s_k[k])
        v_comp[1]['ln_res'] = v_comp[1]['q']*(1-soma)
        v_comp[1]['ln'] = np.exp(v_comp[1]['ln_res'] + v_comp[1]['ln_comb'])
        dict_T[T].append(v_comp[1]['ln'])
        TE[string_dict][T].append(v_comp[1]['ln'])
"""