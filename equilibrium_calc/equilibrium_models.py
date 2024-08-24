import openpyxl
from openpyxl.styles import Font, PatternFill
import math
import os
from py4j.java_gateway import JavaGateway
from abc import ABC


class GatewayJCosmo():
    """
    This class creates mechanism for connecting to the JCOSMO base
    """
    def __init__(self, selected_model):
        self.gateway = JavaGateway(auto_field=True)
        self.JCOSMO = self.gateway.entry_point
        self.model = self.JCOSMO.newModel(selected_model)


class _SLE_GAMMA(ABC):
    def __init__(self, selected_model, solute):
        # model parameters for connecting to JSCOMO
        self._gatewayJCosmo = GatewayJCosmo(selected_model=selected_model)
        self._gateway = self._gatewayJCosmo.gateway
        self._model = self._gatewayJCosmo.model

        # parameters for class utilisation
        self._solute = solute
        self.dict_data = {}

        # parameters for saving in excel file
        self.font = Font(name='Arial',
                         size=12,
                         bold=False,
                         italic=False,
                         underline='none',
                         color='1E211E', )
        self.fill = PatternFill(fill_type='solid',
                                fgColor='6AA7E0', )

    def pop_data_dictionary(self, SLE):
        """
        This method is utilised for populating the data dictionary
        """
        # The excel file must be written like CARBAMAZEPINE example in the folder "input"
        if SLE:
            wb = openpyxl.load_workbook(
                (os.path.dirname(os.path.abspath(__file__)) + f'\\input\\SLE\\{self._solute}.xlsx'))
        else:
            wb = openpyxl.load_workbook(
                (os.path.dirname(os.path.abspath(__file__)) + f'\\input\\GAMMA_INF\\{self._solute}.xlsx'))

        sheet = wb.active
        data = 0
        for l in filter(None, sheet.iter_rows(min_row=2, values_only=True)):
            if not all(value is None for value in l):
                data += 1
                if not f'data_{data}' in self.dict_data:
                    self.dict_data[f'data_{data}'] = {}
                self.dict_data[f'data_{data}']['T'] = l[0]
                self.dict_data[f'data_{data}']['Solvent'] = l[1].split(';')
                self.dict_data[f'data_{data}']['Molar_ratio'] = list(
                    map(float, str(l[2]).replace(',', '.').split(';')))
                if SLE:
                    self.dict_data[f'data_{data}']['x_exp'] = round(l[3], 6)

    def calc_GAMMA_INF(self):
        """
        Method utilised for calculating the SLE, it adds fraction molar and activity coefficient to data_dict
        This method calculates solute's molar fractions by considerating the ln(gamma_inf)
        """

        # tolerance for the iter calculation
        for data, d_values in self.dict_data.items():
            #
            ncomps = 1 + len(d_values['Solvent'])
            comps = self._gateway.new_array(self._gateway.jvm.java.lang.String, ncomps)
            x = self._gateway.new_array(self._gateway.jvm.double, ncomps)

            comps[0] = self._solute
            for i, solvent in enumerate(d_values['Solvent']):
                comps[i + 1] = solvent

            # model
            self._model.setCompounds(comps)
            self._model.setTemperature(d_values['T'])

            # this method obtains solute's fraction ratio by calculating ln_gamma_inf
            x[0] = 0.0
            for i, molar_ratio in enumerate(d_values['Molar_ratio']):
                x[i + 1] = molar_ratio
            self._model.setComposition(x)
            lnGamma = self._model.activityCoefficientLn()
            self.dict_data[data]['ln_gamma_inf'] = lnGamma[0]

    def save_file(self, SLE):
        # creating the excel file
        wb = openpyxl.Workbook()
        ws = wb.active

        # writing on excel file
        ws['A1'] = 'Temperature'
        ws['B1'] = 'Solvent'
        ws['C1'] = 'Molar_ratio'
        if SLE:
            _alf_list = [chr(ord('A') + i) for i in range(8)]
            ws['D1'] = 'x_exp'
            ws['E1'] = 'ln|γ_∞|'
            ws['F1'] = 'x_est_∞'
            ws['G1'] = 'ln|γ|'
            ws['H1'] = 'x_est'
        else:
            _alf_list = [chr(ord('A') + i) for i in range(4)]
            ws['D1'] = 'ln_g_inf'
        for letter in _alf_list:
            ws[f'{letter}1'].font = self.font

        # writing in excel file cells
        k = 2
        for data, d_values in self.dict_data.items():
            ws[f'A{k}'] = d_values['T']
            ws[f'B{k}'] = ';'.join(d_values['Solvent'])
            ws[f'C{k}'] = ";".join(map(str, d_values['Molar_ratio']))
            if SLE:
                ws[f'D{k}'] = d_values['x_exp']
                ws[f'E{k}'] = d_values['ln_gamma_inf']
                ws[f'F{k}'] = d_values['x_est_inf']
                ws[f'G{k}'] = d_values['ln_gamma_est']
                ws[f'H{k}'] = d_values['x_est']
            else:
                ws[f'D{k}'] = d_values['ln_gamma_inf']
            k += 1

        # save the file
        if SLE:
            wb.save((os.path.dirname(os.path.abspath(__file__)) + f'\\output\\SLE\\{self._solute}.xlsx'))
        else:
            wb.save((os.path.dirname(os.path.abspath(__file__)) + f'\\output\\GAMMA_INF\\{self._solute}.xlsx'))


class SLE(_SLE_GAMMA):
    def __init__(self, selected_model, solute, d_H, Tm):
        super().__init__(selected_model, solute)
        # parameters for calculating de equilibrium
        self._d_H = d_H
        self._Tm = Tm
        self._R = 8.314

        self.pop_data_dictionary(SLE=True)
        self.calc_GAMMA_INF()
        self.calc_SLE()
        self.save_file(SLE=True)

    def calc_SLE(self):
        """
        Method utilised for calculating the SLE, it adds fraction molar and activity coefficient to data_dict
        There are two ways to obtain solute's molar fractions, first one is considerating the gamma_inf, and the second
        one is using iterative calculation to resolv the SLE equation (more accurate)
        """
        # tolerance for the iter calculation
        tol = pow(10, -10)
        for data, d_values in self.dict_data.items():
            err = 1
            # SLE constant
            sle_cte = (self._d_H / self._R) * ((1 / self._Tm) - (1 / d_values['T']))
            #
            ncomps = 1 + len(d_values['Solvent'])
            comps = self._gateway.new_array(self._gateway.jvm.java.lang.String, ncomps)
            x = self._gateway.new_array(self._gateway.jvm.double, ncomps)

            comps[0] = self._solute
            for i, solvent in enumerate(d_values['Solvent']):
                comps[i + 1] = solvent

            # set compounds and temperature
            self._model.setCompounds(comps)
            self._model.setTemperature(d_values['T'])

            # this method obtains solute's fraction ratio by calculating ln_gamma_inf
            self.dict_data[data]['x_est_inf'] = math.exp(sle_cte - self.dict_data[data]['ln_gamma_inf'])

            # this method obtains solute's molar fraction by iteraction calculating (more accurate)
            x[0] = d_values['x_exp']  # initial guess for the molar fraction
            while abs(err) > tol:
                for i, molar_ratio in enumerate(d_values['Molar_ratio']):
                    x[i + 1] = molar_ratio * (1 - x[0])

                self._model.setComposition(x)
                lnGamma = self._model.activityCoefficientLn()

                x_calc = math.exp(sle_cte - lnGamma[0])
                err = x_calc - x[0]
                x[0] = x[0] + 0.2*err

            self.dict_data[data]['ln_gamma_est'] = lnGamma[0]
            self.dict_data[data]['x_est'] = math.exp(sle_cte - lnGamma[0])

class Gamma_INF(_SLE_GAMMA):
    def __init__(self, selected_model, solute):
        super().__init__(selected_model, solute)

        self.pop_data_dictionary(SLE=False)
        self.calc_GAMMA_INF()
        self.save_file(SLE=False)


# THE CLASS BELOW IS USED TO CALCULATE LIQUID-LIQUID EXTRACTION
class Partition_K():
    def __init__(self, selected_model, solutes, solvents, temperature, xlsx_name):
        self._selected_model = selected_model
        self._solutes = solutes
        self._solvents = solvents # recebe como [['Name', Total Concentration, Density, MW], ['1-OCTANOL', 8.37, 0.824, 130.2279]]
        self._dict_solvents = {}
        self._temperature = temperature
        self._molar_list = [0.0, 1.0]
        self.dict_lnGamma = {}
        self.dict_K = {}
        self._gatewayJCosmo = GatewayJCosmo(selected_model=self._selected_model)
        self._gateway = self._gatewayJCosmo.gateway
        self._model = self._gatewayJCosmo.model

        self.dict_solv_constructor()
        self.create_dict_lnGamma()
        self.create_dict_K()

        # parameters for saving in excel file
        self._font_st = Font(name='Arial',
                             size=12,
                             bold=False,
                             italic=False,
                             underline='none',
                             color='1E211E', )
        self._fill = PatternFill(fill_type='solid',
                                 fgColor='6AA7E0', )
        self._xlsx_name = xlsx_name
        self.save_file()

    def dict_solv_constructor(self):
        i = 1
        for solvent in self._solvents:
            self._dict_solvents[f'{i}'] = {}
            self._dict_solvents[f'{i}']['Name'] = solvent[0]
            self._dict_solvents[f'{i}']['Total concentration'] = solvent[1]
            self._dict_solvents[f'{i}']['Density'] = solvent[2]
            self._dict_solvents[f'{i}']['MW'] = solvent[3]
            i += 1

    def create_dict_lnGamma(self):
        comps = self._gateway.new_array(self._gateway.jvm.java.lang.String, 2)
        x = self._gateway.new_array(self._gateway.jvm.double, 2)

        for key, value in self._dict_solvents.items():
            comps[1] = value['Name']

            for solute in self._solutes:
                if not solute in self.dict_lnGamma:
                    self.dict_lnGamma[solute] = {}
                comps[0] = solute

                x[0] = self._molar_list[0]
                x[1] = self._molar_list[1]

                self._model.setCompounds(comps)
                self._model.setComposition(x)
                self._model.setTemperature(self._temperature)
                lnGamma = self._model.activityCoefficientLn()
                self.dict_lnGamma[solute][value['Name']] = math.exp(lnGamma[0])

    def create_dict_K(self):
        solvent_1 = self._dict_solvents['1']['Name']
        solvent_2 = self._dict_solvents['2']['Name']

        C_C_1 = self._dict_solvents['2']['Total concentration'] / self._dict_solvents['1']['Total concentration']
        C_C_2 = (self._dict_solvents['2']['Density'] / self._dict_solvents['2']['MW']) / (
                    self._dict_solvents['1']['Density'] / self._dict_solvents['1']['MW'])

        for solute, value in self.dict_lnGamma.items():
            if not solute in self.dict_K:
                self.dict_K[solute] = {}
            self.dict_K[solute]['K_1'] = math.log10(C_C_1*(value[solvent_1] / value[solvent_2]))
            self.dict_K[solute]['K_2'] = math.log10(C_C_2*(value[solvent_1] / value[solvent_2]))
            self.dict_K[solute]['K_3'] = math.log10(value[solvent_1] / value[solvent_2])

    def save_file(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws['A1'] = 'Compound'
        ws['B1'] = 'log|K| = log|(Corg*γ∞water)/(Cwater*γ∞org)|'
        ws['C1'] = 'log|K| = log|CC_rough*(γ∞water/γ∞org)|'
        ws['D1'] = 'log|K| = log|(γ∞org/gγ∞water)|'
        ws['A1'].font = self._font_st
        ws['B1'].font = self._font_st
        ws['C1'].font = self._font_st
        ws['D1'].font = self._font_st

        k = 2
        for solute, value_K in self.dict_K.items():
            ws[f'A{k}'] = solute
            ws[f'B{k}'] = value_K['K_1']
            ws[f'C{k}'] = value_K['K_2']
            ws[f'D{k}'] = value_K['K_3']
            k += 1
        wb.save((os.path.dirname(os.path.abspath(__file__)) + f'\\output\\PARTITION_K\\{self._xlsx_name}.xlsx'))